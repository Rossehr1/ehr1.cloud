#!/usr/bin/env python3
"""
Load EP PAID workbook into supplemental_ep_paid with master NPI gate; then rebuild
merged_ep_paid_npi (latest supplemental row per NPI) unless --skip-merge.

Rows failing the gate go to archive_supplemental_row (full payload JSON). Exact duplicate
(NPI + canonical payload) within the import batch are skipped and counted.

Requires core_npi_provider populated first. Reads DB settings from deploy config (see --config).
Does not read or modify files under Data Originals/ except opening the workbook path you pass.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pymysql
from openpyxl import load_workbook

_TOOLS = Path(__file__).resolve().parent
if str(_TOOLS) not in sys.path:
    sys.path.insert(0, str(_TOOLS))

from load_master_dataset import (  # noqa: E402
    _batch_insert,
    _create_batch,
    _finish_batch,
    _normalize_npi,
    _read_php_config,
)

from merge_ep_paid_to_npi import rebuild_merged_ep_paid  # noqa: E402

DEFAULT_WORKBOOK = Path("Data Originals") / "EP PAID Complete - Final 4-10-26.xlsx"
DEFAULT_CONFIG = Path("deploy/ehr1-cloud-app/includes/config.local.php")


def _unique_headers(raw: list[Any]) -> list[str]:
    """Excel allows duplicate header labels; make JSON keys unique with __idx suffix."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, h in enumerate(raw):
        base = str(h).strip() if h is not None else ""
        if not base:
            base = f"_col{i}"
        n = seen.get(base, 0) + 1
        seen[base] = n
        out.append(base if n == 1 else f"{base}__{i}")
    return out


def _jsonable_cell(value: Any) -> Any:
    """Normalize cell for JSON; keep postal/NPI-adjacent numeric cells as strings."""
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))
        return str(value).strip() or None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, Decimal):
        s = format(value, "f").rstrip("0").rstrip(".")
        return s if s else None
    if isinstance(value, datetime):
        try:
            return value.date().isoformat()
        except Exception:
            return str(value).strip() or None
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    return s if s else None


def _row_payload(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for i, h in enumerate(headers):
        if i >= len(values):
            break
        v = _jsonable_cell(values[i])
        if v is not None:
            payload[h] = v
    return payload


def _canonical_json(payload: dict[str, Any]) -> str:
    return json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))


def _fetch_master_npis(conn: Any) -> set[str]:
    with conn.cursor() as cur:
        cur.execute("SELECT npi FROM core_npi_provider")
        return {row[0] for row in cur.fetchall()}


def _flush_archive(
    cur: Any,
    pending: list[tuple[Any, ...]],
    cols: list[str],
) -> None:
    if pending:
        _batch_insert(cur, "archive_supplemental_row", cols, pending)


def load_ep_paid(
    conn: Any,
    path: Path,
    *,
    batch_size: int = 1000,
    sheet: str | None = None,
    max_rows: int | None = None,
    skip_merge: bool = False,
) -> None:
    master = _fetch_master_npis(conn)
    if not master:
        print("warning: core_npi_provider is empty — all gated rows will archive as NPI_NOT_IN_MASTER", flush=True)

    with conn.cursor() as cur:
        batch_id = _create_batch(
            cur,
            "ep_paid",
            path,
            "EP PAID supplemental payload load (NPI gate + exact row dedupe within batch)",
        )
    conn.commit()

    archive_cols = [
        "source_batch_id",
        "source_file_name",
        "npi_raw",
        "reject_reason",
        "reject_detail",
        "source_line_number",
        "payload_json",
    ]
    insert_cols = ["npi", "payload_json", "source_batch_id"]

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    except KeyError:
        wb.close()
        raise SystemExit(f"Sheet not found: {sheet!r}; available: {wb.sheetnames}")

    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        wb.close()
        raise SystemExit("Workbook has no rows")

    raw_headers = list(header_row)
    try:
        npi_col = next(
            i
            for i, h in enumerate(raw_headers)
            if h is not None and str(h).strip().upper() == "NPI"
        )
    except StopIteration:
        wb.close()
        raise SystemExit("No NPI column in header row")

    headers = _unique_headers(raw_headers)

    loaded = 0
    invalid_npi = 0
    duplicate_skipped = 0
    seen_sig: set[tuple[str, str]] = set()
    pending_ins: list[tuple[Any, ...]] = []
    pending_arc: list[tuple[Any, ...]] = []

    excel_row = 1
    for values in it:
        excel_row += 1
        if max_rows is not None and excel_row - 1 > max_rows:
            break
        raw_npi = None
        if values and npi_col < len(values):
            raw_cell = values[npi_col]
            if raw_cell is not None:
                raw_npi = str(raw_cell).strip()

        payload = _row_payload(headers, values if values else ())
        payload_json_str = _canonical_json(payload)
        digest = hashlib.sha256(payload_json_str.encode("utf-8")).hexdigest()

        npi = _normalize_npi(raw_npi)
        reason: str | None = None
        detail: str | None = None

        if raw_npi is None or (isinstance(raw_npi, str) and raw_npi.strip() == ""):
            reason = "NPI_MISSING"
        elif npi is None:
            reason = "INVALID_NPI_FORMAT"
            detail = raw_npi[:120] if raw_npi else None
        elif npi not in master:
            reason = "NPI_NOT_IN_MASTER"

        if reason:
            invalid_npi += 1
            pending_arc.append(
                (
                    batch_id,
                    path.name,
                    raw_npi[:32] if raw_npi else None,
                    reason,
                    detail,
                    excel_row,
                    payload_json_str if payload else "{}",
                )
            )
        else:
            sig = (npi, digest)
            if sig in seen_sig:
                duplicate_skipped += 1
            else:
                seen_sig.add(sig)
                pending_ins.append((npi, payload_json_str, batch_id))
                loaded += 1

        if len(pending_ins) >= batch_size or len(pending_arc) >= batch_size:
            with conn.cursor() as cur:
                if pending_ins:
                    _batch_insert(cur, "supplemental_ep_paid", insert_cols, pending_ins, upsert=True)
                if pending_arc:
                    _flush_archive(cur, pending_arc, archive_cols)
            conn.commit()
            pending_ins = []
            pending_arc = []
            print(f"{path.name}: committed active {loaded:,} rows (excel row ~{excel_row:,})", flush=True)

    if pending_ins or pending_arc:
        with conn.cursor() as cur:
            if pending_ins:
                _batch_insert(cur, "supplemental_ep_paid", insert_cols, pending_ins, upsert=True)
            _flush_archive(cur, pending_arc, archive_cols)
        conn.commit()

    wb.close()

    with conn.cursor() as cur:
        _finish_batch(cur, batch_id, loaded, invalid_npi, duplicate_skipped)
    conn.commit()

    print(
        f"{path.name}: done — active loaded {loaded:,}, archived (gate) {invalid_npi:,}, "
        f"duplicate_skipped {duplicate_skipped:,}",
        flush=True,
    )

    if skip_merge:
        print("merged_ep_paid_npi: skipped (--skip-merge)", flush=True)
        return
    try:
        n = rebuild_merged_ep_paid(conn)
        conn.commit()
        print(f"merged_ep_paid_npi: rebuilt, {n:,} row(s) (latest supplemental per NPI)", flush=True)
    except Exception as exc:
        conn.rollback()
        print(
            f"warning: merged_ep_paid_npi rebuild failed ({exc!r}) — apply sql/mysql/11_merged_ep_paid.sql re-run "
            f"python tools/merge_ep_paid_to_npi.py",
            flush=True,
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Load EP PAID xlsx into supplemental_ep_paid")
    parser.add_argument("--file", default=str(DEFAULT_WORKBOOK), help="Path to EP PAID xlsx")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    parser.add_argument("--host", default=None)
    parser.add_argument("--port", type=int, default=None)
    parser.add_argument("--batch-size", type=int, default=1000)
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    parser.add_argument("--max-rows", type=int, default=None, help="Stop after N data rows (testing)")
    parser.add_argument("--skip-merge", action="store_true", help="Skip merged_ep_paid_npi rebuild after load")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.is_file():
        raise SystemExit(f"Not a file: {path}")

    cfg = _read_php_config(Path(args.config))
    if args.host is not None:
        cfg["host"] = args.host
    if args.port is not None:
        cfg["port"] = args.port

    conn = pymysql.connect(
        host=cfg["host"],
        port=int(cfg["port"]),
        user=cfg["user"],
        password=cfg["password"],
        database=cfg["database"],
        charset=cfg["charset"],
        autocommit=False,
    )
    try:
        load_ep_paid(
            conn,
            path,
            batch_size=args.batch_size,
            sheet=args.sheet,
            max_rows=args.max_rows,
            skip_merge=args.skip_merge,
        )
    finally:
        conn.close()


if __name__ == "__main__":
    main()
