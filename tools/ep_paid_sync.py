#!/usr/bin/env python3
"""
Load EP PAID Complete *.xlsx into supplemental_ep_paid (payload_json by NPI) and
refresh ep_paid_column_manifest in MySQL (full column list for the Data explorer) and
ep_paid_headers.generated.json for local deploys.

- Reads every worksheet; merges all columns across sheets (ordered union for columns output).
- Merges row payloads by NPI across sheets (later sheet overwrites duplicate keys).
- Empty Excel headers become Column_1, Column_2, …; duplicate names become "Name (2)", etc.

Usage:
  pip install -r tools/requirements-ep-paid.txt
  python tools/ep_paid_sync.py columns --xlsx "Data Originals/EP PAID Complete - Final 4-10-26.xlsx"
  python tools/ep_paid_sync.py load --xlsx "..." --mysql-host 127.0.0.1 --mysql-user u --mysql-password p --mysql-database ehr1

Do not edit Data Originals in place; this script only reads the workbook.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

def _default_headers_json_path() -> Path:
    """Repo layout: tools/ep_paid_sync.py; ad-hoc copy next to script uses that folder."""
    here = Path(__file__).resolve().parent
    if here.name == "tools":
        return here.parent / "deploy/ehr1-cloud-app/includes/ep_paid_headers.generated.json"
    return here / "ep_paid_headers.generated.json"


DEFAULT_HEADERS_JSON = _default_headers_json_path()


def _normalize_headers(raw: list) -> list[str]:
    """Blank headers -> Column_n; duplicate names -> Name (2), Name (3), ..."""
    counts: dict[str, int] = {}
    out: list[str] = []
    for i, cell in enumerate(raw):
        s = "" if cell is None else str(cell).strip()
        base = s if s else f"Column_{i + 1}"
        n = counts.get(base, 0)
        counts[base] = n + 1
        if n == 0:
            out.append(base)
        else:
            out.append(f"{base} ({n + 1})")
    return out


def _find_npi_column(headers: list[str]) -> int:
    for i, h in enumerate(headers):
        low = h.lower()
        if low == "npi" or ("npi" in low and "type" not in low):
            return i
    for i, h in enumerate(headers):
        low = h.lower()
        if low.startswith("national provider"):
            return i
    raise ValueError(
        "Could not find NPI column. Expected a column named like 'NPI'. Headers: "
        + ", ".join(repr(h) for h in headers[:30])
        + ("..." if len(headers) > 30 else "")
    )


def _workbook_headers_union(path: Path) -> list[str]:
    """First row from every sheet; ordered union of normalized header names."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    seen: set[str] = set()
    ordered: list[str] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            continue
        for h in _normalize_headers(list(first)):
            if h not in seen:
                seen.add(h)
                ordered.append(h)
    wb.close()
    return ordered


def _cell_to_json(val: object) -> object:
    """JSON-serializable scalars for MySQL (openpyxl returns datetime for date cells)."""
    if val is None:
        return None
    if isinstance(val, float) and val == int(val) and abs(val) < 1e15:
        return int(val)
    if isinstance(val, (datetime, date, time)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return format(val, "f")
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, bool):
        return val
    return val


def _row_to_payload(headers: list[str], row: tuple) -> dict:
    out: dict[str, object] = {}
    for i, h in enumerate(headers):
        val = row[i] if i < len(row) else None
        out[h] = _cell_to_json(val)
    return out


def _normalize_npi(val: object) -> str | None:
    if val is None:
        return None
    s = re.sub(r"\D", "", str(val))
    if len(s) == 10:
        return s
    return None


def _ordered_manifest_headers(path: Path, by_npi: dict[str, dict]) -> list[str]:
    """Sheet row-1 union first, then any top-level keys seen in merged payloads (same order as explorer)."""
    sheet_order = _workbook_headers_union(path)
    seen: set[str] = set(sheet_order)
    ordered: list[str] = list(sheet_order)
    for payload in by_npi.values():
        if not isinstance(payload, dict):
            continue
        for k in payload.keys():
            ks = k if isinstance(k, str) else str(k)
            if ks not in seen:
                seen.add(ks)
                ordered.append(ks)
    return ordered


def _replace_mysql_manifest(cur: object, headers: list[str]) -> None:
    cur.execute("DELETE FROM ep_paid_column_manifest")
    for i, h in enumerate(headers):
        cur.execute(
            "INSERT INTO ep_paid_column_manifest (ordinal, header_name) VALUES (%s, %s)",
            (i, h[:768]),
        )


def _workbook_merge_payloads_by_npi(path: Path) -> tuple[dict[str, dict], int, int]:
    """Returns (npi -> merged payload, sheets_used, rows_read)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    by_npi: dict[str, dict] = {}
    sheets_used = 0
    rows_read = 0
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        headers = _normalize_headers(list(header_row))
        try:
            npi_i = _find_npi_column(headers)
        except ValueError:
            continue
        sheets_used += 1
        for row in rows:
            rows_read += 1
            npi = _normalize_npi(row[npi_i] if npi_i < len(row) else None)
            if not npi:
                continue
            chunk = _row_to_payload(headers, row)
            if npi in by_npi:
                by_npi[npi].update(chunk)
            else:
                by_npi[npi] = chunk
    wb.close()
    return by_npi, sheets_used, rows_read


def cmd_columns(args: argparse.Namespace) -> None:
    path = Path(args.xlsx)
    headers = _workbook_headers_union(path)
    if not headers:
        raise SystemExit("No headers found in any worksheet (empty workbook?)")
    out = Path(args.out)
    _write_headers_json(out, headers)


def _write_headers_json(out: Path, headers: list[str]) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(headers, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {out} ({len(headers)} distinct column names across all sheets)")


def cmd_load(args: argparse.Namespace) -> None:
    try:
        import pymysql
    except ImportError:
        print("Install pymysql: pip install pymysql", file=sys.stderr)
        sys.exit(1)
    path = Path(args.xlsx)
    by_npi, sheets_used, rows_read = _workbook_merge_payloads_by_npi(path)
    if not by_npi:
        raise SystemExit("No rows with a valid 10-digit NPI found (check NPI column on each sheet).")
    conn = pymysql.connect(
        host=args.mysql_host,
        port=args.mysql_port,
        user=args.mysql_user,
        password=args.mysql_password,
        database=args.mysql_database,
        charset="utf8mb4",
    )
    batch_id = args.batch_id
    if batch_id is None:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO ref_source_batch (source_key, file_name, file_effective_date, row_count_loaded, notes) "
                "VALUES ('ep_paid', %s, CURDATE(), 0, 'EP PAID Complete xlsx load')",
                (path.name,),
            )
            batch_id = cur.lastrowid
        conn.commit()
    headers = _ordered_manifest_headers(path, by_npi)
    loaded = 0
    with conn.cursor() as cur:
        sql = (
            "INSERT INTO supplemental_ep_paid (npi, payload_json, source_batch_id) VALUES (%s, %s, %s) "
            "ON DUPLICATE KEY UPDATE payload_json = VALUES(payload_json), source_batch_id = VALUES(source_batch_id)"
        )
        for npi, payload in by_npi.items():
            cur.execute(sql, (npi, json.dumps(payload, ensure_ascii=False), batch_id))
            loaded += 1
        if batch_id is not None:
            cur.execute(
                "UPDATE ref_source_batch SET row_count_loaded = %s WHERE batch_id = %s",
                (loaded, batch_id),
            )
        if not args.no_manifest_db and headers:
            try:
                _replace_mysql_manifest(cur, headers)
            except Exception as e:
                print(
                    "Warning: could not refresh ep_paid_column_manifest (run migration 07 on the DB?): "
                    + str(e),
                    file=sys.stderr,
                )
    conn.commit()
    conn.close()
    if not args.no_out_headers and headers:
        hpath = Path(args.out_headers)
        try:
            _write_headers_json(hpath, headers)
        except OSError as e:
            print(f"Warning: could not write header manifest {hpath}: {e}", file=sys.stderr)
    print(
        f"Loaded {loaded} NPIs (merged across {sheets_used} sheet(s), {rows_read} data rows scanned). "
        "Re-run load after changing the workbook to refresh payload_json."
    )


def main() -> None:
    p = argparse.ArgumentParser(description="EP PAID xlsx → MySQL + header JSON")
    sub = p.add_subparsers(dest="cmd", required=True)

    pc = sub.add_parser("columns", help="Write ep_paid_headers.generated.json from all sheet headers")
    pc.add_argument("--xlsx", required=True, help="Path to EP PAID Complete xlsx")
    pc.add_argument("--out", default=str(DEFAULT_HEADERS_JSON), help="Output JSON path")
    pc.set_defaults(func=cmd_columns)

    pl = sub.add_parser("load", help="Upsert merged payloads into supplemental_ep_paid")
    pl.add_argument("--xlsx", required=True)
    pl.add_argument("--mysql-host", default="127.0.0.1")
    pl.add_argument("--mysql-port", type=int, default=3306)
    pl.add_argument("--mysql-user", required=True)
    pl.add_argument("--mysql-password", default="")
    pl.add_argument("--mysql-database", required=True)
    pl.add_argument(
        "--batch-id",
        type=int,
        default=None,
        help="Existing ref_source_batch.batch_id; default creates new ep_paid batch",
    )
    pl.add_argument(
        "--out-headers",
        default=str(DEFAULT_HEADERS_JSON),
        help="Write ep_paid_headers.generated.json (union of all sheet row-1 headers) for the Data explorer",
    )
    pl.add_argument(
        "--no-out-headers",
        action="store_true",
        help="Skip writing ep_paid_headers.generated.json",
    )
    pl.add_argument(
        "--no-manifest-db",
        action="store_true",
        help="Skip DELETE/INSERT into ep_paid_column_manifest (requires migration 07)",
    )
    pl.set_defaults(func=cmd_load)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
