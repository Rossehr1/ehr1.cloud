#!/usr/bin/env python3
"""Smoke-test EP PAID workbook parsing without MySQL (headers, NPI column, JSON shape)."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

_REPO = Path(__file__).resolve().parents[1]
_TOOLS = _REPO / "tools"
sys.path.insert(0, str(_TOOLS))

from load_ep_paid import (  # noqa: E402
    _canonical_json,
    _row_payload,
    _unique_headers,
)


def main() -> None:
    root = _REPO
    xlsx = root / "Data Originals" / "EP PAID Complete - Final 4-10-26.xlsx"
    if not xlsx.is_file():
        print(f"skip: not found {xlsx}", flush=True)
        return
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    raw_h = list(header)
    npi_col = next(
        i
        for i, h in enumerate(raw_h)
        if h is not None and str(h).strip().upper() == "NPI"
    )
    headers = _unique_headers(raw_h)
    assert headers[npi_col] == "NPI" or str(raw_h[npi_col]).strip().upper() == "NPI"
    row2 = next(it)
    payload = _row_payload(headers, row2)
    js = _canonical_json(payload)
    print("OK sheet:", wb.sheetnames[0], flush=True)
    print("OK header cols:", len(headers), "npi_col:", npi_col, flush=True)
    print("OK sample keys:", sorted(payload.keys())[:12], "...", flush=True)
    print("OK canonical JSON length:", len(js), flush=True)
    wb.close()


if __name__ == "__main__":
    main()
